"""
io_utils/semantic_io.py

Read/write *_PJ_d2m_semantic_<rater>.xlsx files in the exact schema already
in use (verified against 0b6823317195419dbd947019dbf72d96_PJ_d2m_semantic_Saw.xlsx):

  Sheet "Annotations":
    file, pcaso_var, start, stop, Comments, T1, E1, T2, E2, ..., T60, E60
    - file:      the original full-length case video's identifier
    - pcaso_var: stitch id, e.g. "stitch_00.25"
    - start/stop: stitch boundaries, as time-of-day in the ORIGINAL video's
                  timebase (NOT clip-relative -- see CLAUDE.md)
    - Tn/En:     timestamp (original video timebase) + event label, one
                 pair per explicit annotated event, left-to-right in time
                 order. Unused pairs are left blank.

  Sheet "Lists": reference list of event types (used for xlsx data
  validation dropdowns in the original files; we regenerate it identically).

  Sheet "QualityChecks": human-readable rules, reproduced verbatim so the
  file stays self-documenting if someone opens it outside this tool.

All timestamps are Python `datetime.time` objects so openpyxl round-trips
them as native Excel time cells (mm:ss.00-style display), exactly like the
source files.
"""
from __future__ import annotations
from dataclasses import dataclass, field
from datetime import time
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from core.config import EVENT_TYPES, MAX_EVENT_PAIRS

ANNOTATIONS_SHEET = "Annotations"
LISTS_SHEET = "Lists"
QUALITY_SHEET = "QualityChecks"

QUALITY_CHECK_TEXT = [
    "QUALITY CHECKS (key rules)",
    "\u2022 Yank and Path must be integers in {1,2,3,4}.",
    "\u2022 First event must be TIP_ENTRY; no TIP_WITHDRAWAL_NEAR before it.",
    "\u2022 After the first TIP_EXIT_FAR, later TIP_EXIT_FAR must be preceded by TIP_WITHDRAWAL_FAR.",
    "\u2022 TAIL_EXIT only after a TIP_EXIT_FAR; at most one TAIL_EXIT per row.",
    "\u2022 READJUSTMENT_START must end at READJUSTMENT_END or TIP_WITHDRAWAL_NEAR; no overlaps.",
    "\u2022 CAM_REPOSITION_START must be closed by CAM_REPOSITION_END; may overlap other events.",
    "\u2022 Timestamps non-decreasing left\u2192right; duplicates only if simultaneous.",
    "\u2022 Time cells formatted as mm:ss.00.",
]

TIME_NUMBER_FORMAT = "mm:ss.00"


@dataclass
class StitchRow:
    file: str
    pcaso_var: str            # stitch id, e.g. "stitch_00.25"
    start: Optional[time]
    stop: Optional[time]
    comments: Optional[str]
    events: list[tuple[Optional[time], Optional[str]]] = field(default_factory=list)

    def sorted_events(self) -> list[tuple[time, str]]:
        return [(t, e) for (t, e) in self.events if t is not None and e]


def _time_to_seconds(t) -> Optional[float]:
    if t is None:
        return None
    if not isinstance(t, time):
        return None  # malformed cell (e.g. a typo'd string) -- caller reports this
    return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1e6


def seconds_to_time(sec: float) -> time:
    sec = max(0.0, sec)
    h = int(sec // 3600)
    rem = sec - h * 3600
    m = int(rem // 60)
    rem -= m * 60
    s = int(rem)
    micro = int(round((rem - s) * 1e6))
    if micro >= 1_000_000:
        micro -= 1_000_000
        s += 1
    if s >= 60:
        s -= 60
        m += 1
    if m >= 60:
        m -= 60
        h += 1
    return time(hour=min(h, 23), minute=m, second=s, microsecond=micro)


def _new_workbook(bodypart_events: list[str] = EVENT_TYPES) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ANNOTATIONS_SHEET
    header = ["file", "pcaso_var", "start", "stop", "Comments"]
    for i in range(1, MAX_EVENT_PAIRS + 1):
        header += [f"T{i}", f"E{i}"]
    ws.append(header)

    lists_ws = wb.create_sheet(LISTS_SHEET)
    for i, ev in enumerate(bodypart_events):
        lists_ws.append([ev, str(i + 1) if i < 4 else None])

    qc_ws = wb.create_sheet(QUALITY_SHEET)
    for line in QUALITY_CHECK_TEXT:
        qc_ws.append([line])
    return wb


def load_or_create(xlsx_path: Path) -> openpyxl.Workbook:
    if xlsx_path.exists():
        return openpyxl.load_workbook(xlsx_path)
    return _new_workbook()


def read_rows(xlsx_path: Path) -> list[StitchRow]:
    if not xlsx_path.exists():
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if ANNOTATIONS_SHEET not in wb.sheetnames:
        return []
    ws = wb[ANNOTATIONS_SHEET]
    rows_out: list[StitchRow] = []
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        file_, pcaso_var, start, stop, comments = row[0], row[1], row[2], row[3], row[4]
        events: list[tuple[Optional[time], Optional[str]]] = []
        rest = row[5:]
        for i in range(0, len(rest) - 1, 2):
            events.append((rest[i], rest[i + 1]))
        rows_out.append(StitchRow(
            file=file_, pcaso_var=pcaso_var, start=start, stop=stop,
            comments=comments, events=events,
        ))
    return rows_out


def upsert_row(xlsx_path: Path, new_row: StitchRow) -> None:
    """Insert or replace the row matching (file, pcaso_var); leaves every
    other stitch's row in the workbook untouched. This is the operation the
    Semantic tab calls on every save, since one xlsx holds all stitches for
    one rater across a whole case."""
    wb = load_or_create(xlsx_path)
    if ANNOTATIONS_SHEET not in wb.sheetnames:
        # existing file with unexpected sheet naming -- rebuild fresh
        wb = _new_workbook()
    ws = wb[ANNOTATIONS_SHEET]

    target_row_idx: Optional[int] = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == new_row.file and \
           ws.cell(row=r, column=2).value == new_row.pcaso_var:
            target_row_idx = r
            break
    if target_row_idx is None:
        target_row_idx = ws.max_row + 1 if ws.cell(row=ws.max_row, column=1).value else ws.max_row

    events = new_row.sorted_events()
    if len(events) > MAX_EVENT_PAIRS:
        raise ValueError(
            f"{new_row.pcaso_var}: {len(events)} events exceeds MAX_EVENT_PAIRS={MAX_EVENT_PAIRS}"
        )

    ws.cell(row=target_row_idx, column=1, value=new_row.file)
    ws.cell(row=target_row_idx, column=2, value=new_row.pcaso_var)
    c_start = ws.cell(row=target_row_idx, column=3, value=new_row.start)
    c_stop = ws.cell(row=target_row_idx, column=4, value=new_row.stop)
    c_start.number_format = TIME_NUMBER_FORMAT
    c_stop.number_format = TIME_NUMBER_FORMAT
    ws.cell(row=target_row_idx, column=5, value=new_row.comments)

    col = 6
    for i in range(MAX_EVENT_PAIRS):
        t, ev = events[i] if i < len(events) else (None, None)
        tc = ws.cell(row=target_row_idx, column=col, value=t)
        if t is not None:
            tc.number_format = TIME_NUMBER_FORMAT
        ws.cell(row=target_row_idx, column=col + 1, value=ev)
        col += 2

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Quality-check validator -- mirrors the QUALITY_CHECK_TEXT rules exactly.
# ---------------------------------------------------------------------------
def validate_events(events: list[tuple[time, str]]) -> list[str]:
    """Returns a list of human-readable problems; empty list == clean."""
    problems: list[str] = []
    if not events:
        return problems

    for i, (t, _lbl) in enumerate(events):
        if not isinstance(t, time):
            problems.append(
                f"Event {i+1}: timestamp cell is not a valid time value ({t!r}). "
                f"This usually means a typo (e.g. '13.19.73' instead of a time) -- fix the cell."
            )

    secs = [_time_to_seconds(t) for t, _ in events]
    labels = [e for _, e in events]

    # non-decreasing left-to-right
    for i in range(1, len(secs)):
        if secs[i] is not None and secs[i - 1] is not None and secs[i] < secs[i - 1]:
            problems.append(
                f"Event {i+1} ({labels[i]} @ {events[i][0]}) is earlier than "
                f"event {i} ({labels[i-1]} @ {events[i-1][0]}) -- timestamps must be non-decreasing."
            )

    # First *needle-interaction* event must be TIP_ENTRY; no TIP_WITHDRAWAL_NEAR
    # before it. CAM_REPOSITION_START/END are exempt from ordering checks --
    # per the rules they "may overlap other events" and routinely occur
    # before the first TIP_ENTRY (e.g. framing the shot).
    non_cam_idx = [i for i, l in enumerate(labels) if not l.startswith("CAM_REPOSITION")]
    first_tip_entry_idx = next((i for i, l in enumerate(labels) if l == "TIP_ENTRY"), None)
    if non_cam_idx and labels[non_cam_idx[0]] != "TIP_ENTRY":
        i0 = non_cam_idx[0]
        problems.append(
            f"First needle-interaction event is '{labels[i0]}' (event {i0+1}), "
            f"but it must be TIP_ENTRY."
        )
    if first_tip_entry_idx is not None:
        for i in range(first_tip_entry_idx):
            if labels[i] == "TIP_WITHDRAWAL_NEAR":
                problems.append(f"Event {i+1}: TIP_WITHDRAWAL_NEAR occurs before the first TIP_ENTRY.")

    # After the first TIP_EXIT_FAR *within a suturing pass*, a later
    # TIP_EXIT_FAR in that same pass must be preceded by TIP_WITHDRAWAL_FAR.
    # Each TIP_ENTRY starts a new pass and resets this check -- otherwise
    # every stitch with more than one needle pass would be flagged.
    seen_first_exit_far = False
    since_last_exit_far_had_withdrawal_far = False
    for i, l in enumerate(labels):
        if l == "TIP_ENTRY":
            seen_first_exit_far = False
            since_last_exit_far_had_withdrawal_far = False
        elif l == "TIP_EXIT_FAR":
            if seen_first_exit_far and not since_last_exit_far_had_withdrawal_far:
                problems.append(
                    f"Event {i+1}: TIP_EXIT_FAR not preceded by TIP_WITHDRAWAL_FAR "
                    f"since the previous TIP_EXIT_FAR in this pass."
                )
            seen_first_exit_far = True
            since_last_exit_far_had_withdrawal_far = False
        elif l == "TIP_WITHDRAWAL_FAR":
            since_last_exit_far_had_withdrawal_far = True

    # TAIL_EXIT only after a TIP_EXIT_FAR; at most one TAIL_EXIT per "pass"
    # (we check: at most one TAIL_EXIT between consecutive TIP_ENTRYs)
    tail_exit_count_since_entry = 0
    exit_far_seen_since_entry = False
    for i, l in enumerate(labels):
        if l == "TIP_ENTRY":
            tail_exit_count_since_entry = 0
            exit_far_seen_since_entry = False
        elif l == "TIP_EXIT_FAR":
            exit_far_seen_since_entry = True
        elif l == "TAIL_EXIT":
            if not exit_far_seen_since_entry:
                problems.append(f"Event {i+1}: TAIL_EXIT occurs without a preceding TIP_EXIT_FAR.")
            tail_exit_count_since_entry += 1
            if tail_exit_count_since_entry > 1:
                problems.append(f"Event {i+1}: more than one TAIL_EXIT since the last TIP_ENTRY.")

    # READJUSTMENT_START must end at READJUSTMENT_END or TIP_WITHDRAWAL_NEAR; no overlaps
    open_readj: Optional[int] = None
    for i, l in enumerate(labels):
        if l == "READJUSTMENT_START":
            if open_readj is not None:
                problems.append(f"Event {i+1}: READJUSTMENT_START opened while another is still open.")
            open_readj = i
        elif l in ("READJUSTMENT_END", "TIP_WITHDRAWAL_NEAR"):
            if l == "READJUSTMENT_END":
                if open_readj is None:
                    problems.append(f"Event {i+1}: READJUSTMENT_END with no matching READJUSTMENT_START.")
                else:
                    open_readj = None
    if open_readj is not None:
        problems.append(
            f"Event {open_readj+1}: READJUSTMENT_START is never closed by "
            f"READJUSTMENT_END or TIP_WITHDRAWAL_NEAR."
        )

    # CAM_REPOSITION_START must be closed by CAM_REPOSITION_END (overlaps OK)
    open_cam = 0
    for i, l in enumerate(labels):
        if l == "CAM_REPOSITION_START":
            open_cam += 1
        elif l == "CAM_REPOSITION_END":
            open_cam -= 1
            if open_cam < 0:
                problems.append(f"Event {i+1}: CAM_REPOSITION_END with no matching CAM_REPOSITION_START.")
                open_cam = 0
    if open_cam > 0:
        problems.append(f"{open_cam} CAM_REPOSITION_START event(s) never closed by CAM_REPOSITION_END.")

    return problems
